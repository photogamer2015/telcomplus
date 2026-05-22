from pathlib import Path

from django.contrib.auth.models import Group, User
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from parque.models import Equipo
from parque.qr import actualizar_qr


class Command(BaseCommand):
    help = 'Carga roles, usuarios demo y equipos desde el Excel base.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel',
            default='data/equipos_demo.xlsx',
            help='Ruta del archivo Excel con la hoja BD PARQUE.',
        )
        parser.add_argument(
            '--base-url',
            default='http://127.0.0.1:8000',
            help='URL base que quedara codificada en los QR.',
        )

    def handle(self, *args, **options):
        self._crear_roles_y_usuarios()
        self._cargar_equipos(Path(options['excel']), options['base_url'])

    def _crear_roles_y_usuarios(self):
        usuarios = [
            ('admin', 'Admin', True),
            ('tecnico', 'Tecnico', False),
            ('contrato', 'Admin de contrato', False),
        ]
        for username, group_name, is_superuser in usuarios:
            group, _ = Group.objects.get_or_create(name=group_name)
            user, created = User.objects.get_or_create(username=username, defaults={
                'is_staff': True,
                'is_superuser': is_superuser,
            })
            if created:
                user.set_password(f'{username}123')
            user.is_staff = True
            user.is_superuser = is_superuser
            user.save()
            user.groups.add(group)
            self.stdout.write(self.style.SUCCESS(f'Usuario listo: {username} / {username}123'))

    def _cargar_equipos(self, excel_path, base_url):
        workbook = load_workbook(excel_path, data_only=True)
        worksheet = workbook['BD PARQUE']
        headers = [cell.value for cell in worksheet[1]]
        index = {name: pos for pos, name in enumerate(headers) if name}
        creados = 0
        actualizados = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row[index['Serie']]:
                continue
            tipo = str(row[index['Tipo de equipo']] or 'OTRO').strip().upper()
            if tipo not in dict(Equipo.TIPO_CHOICES):
                tipo = Equipo.OTRO
            data = {
                'tipo': tipo,
                'marca': str(row[index['Marca']] or '').strip(),
                'modelo': str(row[index['Modelo']] or '').strip(),
                'empresa_atencion': str(row[index['Empresa de atención']] or '').strip(),
                'razon_social_cliente': str(row[index['Razon Social Cliente']] or '').strip(),
                'provincia': str(row[index['Provincia']] or '').strip(),
                'ciudad': str(row[index['Ciudad']] or '').strip(),
                'direccion': str(row[index['Dirección']] or '').strip(),
                'area_departamento': str(row[index['Área/Departamento']] or '').strip(),
                'ubicacion_coordenadas': str(row[index['Ubicación/coordenadas']] or '').strip(),
            }
            equipo, created = Equipo.objects.update_or_create(
                serie=str(row[index['Serie']]).strip(),
                defaults=data,
            )
            actualizar_qr(equipo, base_url)
            creados += int(created)
            actualizados += int(not created)

        self.stdout.write(self.style.SUCCESS(f'Equipos creados: {creados}; actualizados: {actualizados}'))
